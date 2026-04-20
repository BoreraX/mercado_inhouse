import csv
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


REQUIRED_COLUMNS = [
    "Codigo de Barras",
    "Nome do Produto",
    "Status",
    "Preço Final (consumidor)",
    "Qtd.Estoque",
    "Data de vencimento",
]


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"C:\Users\JP_Desenvolvimento\Downloads\planogram-2026-03-02.xlsx"
    )
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/produtos.csv")

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    worksheet, headers = find_product_sheet(workbook)
    header_index = {header: index for index, header in enumerate(headers)}

    rows = []
    status_counts = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(has_value(value) for value in row):
            continue

        item = {}
        for column in REQUIRED_COLUMNS:
            value = row[header_index[column]] if header_index[column] < len(row) else ""
            item[column] = format_value(column, value)

        rows.append(item)
        status = item["Status"]
        status_counts[status] = status_counts.get(status, 0) + 1

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=REQUIRED_COLUMNS, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)

    print(f"Arquivo gerado: {output.resolve()}")
    print(f"Produtos exportados: {len(rows)}")
    print(f"Status: {status_counts}")


def find_product_sheet(workbook):
    for worksheet in workbook.worksheets:
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(value).strip() if value is not None else "" for value in first_row or []]

        if all(column in headers for column in REQUIRED_COLUMNS):
            return worksheet, headers

    raise ValueError("Nenhuma aba com as colunas esperadas foi encontrada.")


def has_value(value):
    return value is not None and str(value).strip() != ""


def format_value(column, value):
    if value is None:
        return ""

    if column == "Preço Final (consumidor)":
        number = parse_number(value)
        return f"R$ {number:.2f}".replace(".", ",")

    if column == "Data de vencimento" and isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def parse_number(value):
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text and text.rfind(",") > text.rfind("."):
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")

    return float(text or 0)


if __name__ == "__main__":
    main()
